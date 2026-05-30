from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from .models import Booking, SystemConfiguration
from .forms import BookingForm, Esp32WifiConfigForm, AdminPasswordChangeForm
from .utils import (
    print_to_hardware,
    check_printer_status,
    scan_esp32_wifi_networks,
    connect_esp32_wifi,
    discover_esp32_on_local_network,
)
import pandas as pd
import plotly.express as px
import plotly.io as pio
from datetime import timedelta, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io


def create_multi_day_bookings(form_data):
    try:
        booking_type   = form_data.get('booking_type')
        start_date     = form_data.get('date_of_event')
        days_to_book   = form_data.get('days_to_book', [])
        start_time     = form_data.get('start_time')
        duration_hours = form_data.get('duration_hours')

        if not all([booking_type, start_date, start_time, duration_hours]):
            raise ValueError("Missing required booking information")

        if booking_type == 'SINGLE':
            booking = Booking(
                event_name=form_data.get('event_name'),
                customer_name=form_data.get('customer_name'),
                contact_number=form_data.get('contact_number'),
                date_of_event=start_date,
                start_time=start_time,
                duration_hours=duration_hours,
                booking_type='SINGLE',
                parent_booking=None,
                payment_status=form_data.get('payment_status', 'UNPAID'),
                notes=form_data.get('notes', ''),
            )
            booking.save()
            return booking

        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        days_int          = [int(d) for d in days_to_book]
        days_until_monday = start_date.weekday()
        actual_monday     = start_date - timedelta(days=days_until_monday)

        parent_booking = Booking(
            event_name=form_data.get('event_name'),
            customer_name=form_data.get('customer_name'),
            contact_number=form_data.get('contact_number'),
            date_of_event=actual_monday,
            start_time=start_time,
            duration_hours=duration_hours,
            booking_type='MULTIPLE',
            parent_booking=None,
            payment_status=form_data.get('payment_status', 'UNPAID'),
            notes=form_data.get('notes', ''),
        )
        parent_booking.save()

        for day_offset in sorted(days_int):
            booking_date = actual_monday + timedelta(days=day_offset)
            day_booking = Booking(
                event_name=form_data.get('event_name'),
                customer_name=form_data.get('customer_name'),
                contact_number=form_data.get('contact_number'),
                date_of_event=booking_date,
                start_time=start_time,
                duration_hours=duration_hours,
                booking_type='SINGLE',
                parent_booking=parent_booking,
                payment_status=form_data.get('payment_status', 'UNPAID'),
                notes=form_data.get('notes', ''),
            )
            day_booking.save()

        return parent_booking

    except Exception as e:
        raise ValueError(f"Error creating booking: {str(e)}")


@login_required
def dashboard(request):
    today = timezone.now().date()

    active_bookings  = Booking.objects.filter(status='ACTIVE')
    bookings_today   = active_bookings.filter(date_of_event=today).count()
    pending_payments = active_bookings.exclude(payment_status='PAID').count()
    total_cancelled  = Booking.objects.filter(status='CANCELLED').count()

    booking_dates = active_bookings.values('date_of_event')
    chart_html = "<p class='text-muted text-center py-3'>No booking data for chart yet.</p>"
    if booking_dates.exists():
        df = pd.DataFrame(list(booking_dates))
        df['month'] = pd.to_datetime(df['date_of_event']).dt.strftime('%b %Y')
        month_counts = df.groupby('month').size().reset_index(name='count')
        fig = px.bar(
            month_counts, x='month', y='count',
            title='Monthly Booking Trends',
            color_discrete_sequence=['#0d6efd'],
            labels={'month': 'Month', 'count': 'Bookings'},
        )
        fig.update_layout(
            plot_bgcolor='white', paper_bgcolor='white',
            margin=dict(l=20, r=20, t=40, b=20),
            font=dict(family='Inter, sans-serif'),
        )
        chart_html = pio.to_html(fig, full_html=False)

    context = {
        'bookings_today':   bookings_today,
        'pending_payments': pending_payments,
        'total_cancelled':  total_cancelled,
        'chart_html':       chart_html,
        'recent_bookings':  active_bookings.order_by('-created_at')[:5],
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def booking_calendar_data(request):
    bookings = Booking.objects.filter(
        status='ACTIVE'
    ).values(
        'id', 'event_name', 'customer_name',
        'date_of_event', 'start_time', 'status', 'payment_status',
    )
    events = [
        {
            'id':    b['id'],
            'title': f"{b['event_name']} ({b['customer_name']})",
            'start': f"{b['date_of_event']}T{b['start_time']}",
            'color': '#1a56db',
            'url':   f"/bookings/?search={b['id']}",
            'extendedProps': {
                'status':         b['status'],
                'payment_status': b['payment_status'],
                'customer':       b['customer_name'],
            },
        }
        for b in bookings
    ]
    return JsonResponse(events, safe=False)


@login_required
def booking_list(request):
    search_query   = request.GET.get('search', '').strip()
    status_filter  = request.GET.get('status', '').strip()
    payment_filter = request.GET.get('payment', '').strip()
    show_cancelled = request.GET.get('show_cancelled', '0').strip()

    if status_filter:
        bookings = Booking.objects.filter(
            status=status_filter
        ).order_by('-date_of_event')
    elif show_cancelled == '1':
        bookings = Booking.objects.all().order_by('-date_of_event')
    else:
        bookings = Booking.objects.filter(
            status='ACTIVE'
        ).order_by('-date_of_event')

    if search_query:
        bookings = bookings.filter(
            Q(event_name__icontains=search_query)     |
            Q(customer_name__icontains=search_query)  |
            Q(receipt_number__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )

    if payment_filter:
        bookings = bookings.filter(payment_status=payment_filter)

    cancelled_count = Booking.objects.filter(status='CANCELLED').count()
    paginator = Paginator(bookings, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj':         page_obj,
        'search_query':     search_query,
        'status_filter':    status_filter,
        'payment_filter':   payment_filter,
        'show_cancelled':   show_cancelled,
        'cancelled_count':  cancelled_count,
        'total_count':      paginator.count,
        'today':            timezone.now().date(),
        'booking_statuses': Booking.BOOKING_STATUS_CHOICES,
        'payment_statuses': Booking.PAYMENT_STATUS_CHOICES,
    }
    return render(request, 'core/booking_list.html', context)


@login_required
def add_booking(request):
    if request.method == 'POST':
        form = BookingForm(request.POST)
        if form.is_valid():
            try:
                booking = create_multi_day_bookings(form.cleaned_data)
                if booking.booking_type == 'MULTIPLE':
                    children = booking.child_bookings.count()
                    messages.success(
                        request,
                        f'✓ Multi-day booking "{booking.event_name}" for {booking.customer_name} '
                        f'({children} days) added successfully! Receipt: {booking.receipt_number}'
                    )
                else:
                    messages.success(
                        request,
                        f'✓ Booking "{booking.event_name}" for {booking.customer_name} '
                        f'added successfully! Receipt: {booking.receipt_number}'
                    )
                return redirect('booking_list')
            except ValueError as e:
                messages.error(request, f'Booking Error: {str(e)}')
        else:
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, str(error))
    else:
        form = BookingForm()

    return render(request, 'core/booking_form.html', {
        'form':       form,
        'form_title': 'Add New Booking',
        'btn_label':  'Save Booking',
        'is_edit':    False,
    })


@login_required
def edit_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    if booking.is_cancelled:
        messages.warning(request, 'Cancelled bookings cannot be edited.')
        return redirect('booking_list')

    if booking.booking_type == 'MULTIPLE':
        messages.warning(request, 'Multi-day bookings cannot be edited. Please cancel and create a new booking.')
        return redirect('booking_list')

    if request.method == 'POST':
        form = BookingForm(request.POST, instance=booking)
        if form.is_valid():
            form.save()
            messages.success(request, f'✓ Booking "{booking.event_name}" updated successfully.')
            return redirect('booking_list')
        else:
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, str(error))
    else:
        form = BookingForm(instance=booking)

    return render(request, 'core/booking_form.html', {
        'form':       form,
        'form_title': f'Edit Booking — {booking.event_name}',
        'btn_label':  'Update Booking',
        'is_edit':    True,
        'booking':    booking,
    })


@login_required
def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    if booking.status == 'CANCELLED':
        messages.warning(request, f'Booking "{booking.event_name}" is already cancelled.')
        return redirect('booking_list')

    if request.method == 'POST':
        reason = request.POST.get('cancellation_reason', '').strip()
        if not reason:
            messages.error(request, 'Please provide a cancellation reason.')
            return render(request, 'core/cancel_booking.html', {'booking': booking})

        try:
            if booking.booking_type == 'MULTIPLE':
                for child in booking.child_bookings.all():
                    child.status              = 'CANCELLED'
                    child.cancellation_reason = reason
                    child.cancelled_at        = timezone.now()
                    child.cancelled_by        = request.user.get_full_name() or request.user.username
                    child.save()

            booking.status              = 'CANCELLED'
            booking.cancellation_reason = reason
            booking.cancelled_at        = timezone.now()
            booking.cancelled_by        = request.user.get_full_name() or request.user.username
            booking.save()

            if booking.booking_type == 'MULTIPLE':
                messages.success(
                    request,
                    f'✓ Multi-day booking "{booking.event_name}" and all its days have been cancelled.'
                )
            else:
                messages.success(
                    request,
                    f'✓ Booking "{booking.event_name}" for {booking.customer_name} has been cancelled.'
                )
            return redirect('booking_list')

        except Exception as e:
            messages.error(request, f'Error cancelling booking: {str(e)}')
            return render(request, 'core/cancel_booking.html', {'booking': booking})

    return render(request, 'core/cancel_booking.html', {'booking': booking})


@login_required
def print_receipt(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    if booking.is_cancelled:
        messages.error(request, 'Cannot print receipt for a cancelled booking.')
        return redirect('booking_list')
    success, message = print_to_hardware(booking)
    if success:
        messages.success(request, f'✓ Receipt for "{booking.event_name}" sent to printer.')
    else:
        messages.error(request, f'Printing failed: {message}')
    return redirect('booking_list')


@login_required
def esp32_wifi_settings(request):
    config = SystemConfiguration.objects.first()
    if not config:
        config = SystemConfiguration.objects.create()

    form = Esp32WifiConfigForm(request.POST or None, instance=config)
    available_networks = []
    status_online, status_message, status_data = check_printer_status(config)

    if request.method == 'POST':
        if form.is_valid():
            config = form.save()
            if request.POST.get('connect_now'):
                ssid = form.cleaned_data.get('esp32_wifi_ssid')
                password = form.cleaned_data.get('esp32_wifi_password')
                if ssid:
                    connected, message = connect_esp32_wifi(ssid, password, config)
                    if connected:
                        messages.success(request, '✓ Wi-Fi settings saved and ESP32 connect request sent.')
                    else:
                        messages.error(request, f'Wi-Fi connect failed: {message}')
                else:
                    messages.error(request, 'Please select or enter a 2.4GHz SSID before connecting.')
            else:
                messages.success(request, '✓ ESP32 configuration saved.')
            return redirect('esp32_wifi_settings')
        else:
            messages.error(request, 'Please correct the highlighted ESP32 form fields.')

    if request.GET.get('scan') == '1':
        success, result = scan_esp32_wifi_networks(config)
        if success:
            available_networks = result
        else:
            messages.error(request, f'Network scan failed: {result}')

    if request.GET.get('discover') == '1':
        success, result = discover_esp32_on_local_network(config)
        if success:
            config.esp32_ip = result
            config.save()
            messages.success(request, f'ESP32 discovered at {result} and saved to configuration.')
            return redirect('esp32_wifi_settings')
        else:
            messages.error(request, f'Discovery failed: {result}')

    if request.GET.get('refresh') == '1':
        # Re-check device status; if offline, try discovery and save found IP
        status_online, status_message, status_data = check_printer_status(config)
        if status_online:
            messages.success(request, 'ESP32 is online.')
            return redirect('esp32_wifi_settings')
        # try discovery fallback
        success, result = discover_esp32_on_local_network(config)
        if success:
            config.esp32_ip = result
            config.save()
            messages.success(request, f'ESP32 discovered at {result} and saved to configuration.')
        else:
            messages.error(request, f'Refresh failed: {result}')
        return redirect('esp32_wifi_settings')

    return render(request, 'core/esp32_wifi_settings.html', {
        'form': form,
        'status_online': status_online,
        'status_message': status_message,
        'status_data': status_data,
        'available_networks': available_networks,
        'config': config,
    })


@login_required
def delete_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    if booking.status != 'CANCELLED':
        messages.error(request, 'Only cancelled bookings can be deleted.')
        return redirect('booking_list')

    if request.method == 'POST':
        try:
            # If it's a multi-day parent booking, remove its child bookings first
            if booking.booking_type == 'MULTIPLE':
                booking.child_bookings.all().delete()

            booking.delete()
            messages.success(request, f'✓ Cancelled booking "{booking.event_name}" deleted.')
            return redirect('booking_list')
        except Exception as e:
            messages.error(request, f'Error deleting booking: {str(e)}')
            return redirect('booking_list')

    return render(request, 'core/confirm_delete_booking.html', {'booking': booking})


@login_required
def bulk_delete_bookings(request):
    if request.method != 'POST':
        messages.error(request, 'Invalid request method for bulk delete.')
        return redirect('booking_list')

    ids = request.POST.getlist('selected_ids')
    if not ids:
        messages.error(request, 'No bookings selected for deletion.')
        return redirect('/bookings/?show_cancelled=1')

    # Filter only cancelled bookings to avoid accidental deletes
    qs = Booking.objects.filter(id__in=ids, status='CANCELLED')
    to_delete_count = qs.count()

    if to_delete_count == 0:
        messages.error(request, 'Selected bookings cannot be deleted (only cancelled bookings may be removed).')
        return redirect('/bookings/?show_cancelled=1')

    try:
        qs.delete()
        messages.success(request, f'✓ Deleted {to_delete_count} cancelled booking(s).')
    except Exception as e:
        messages.error(request, f'Error deleting bookings: {str(e)}')

    return redirect('/bookings/?show_cancelled=1')


# ══════════════════════════════════════════════════════════════════
#  EXPORT VIEWS
# ══════════════════════════════════════════════════════════════════

def _get_export_queryset(request):
    today     = timezone.now().date()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to',   '').strip()
    if not date_from:
        date_from = today.replace(day=1).strftime('%Y-%m-%d')
    if not date_to:
        date_to = today.strftime('%Y-%m-%d')
    try:
        dt_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        dt_to   = datetime.strptime(date_to,   '%Y-%m-%d').date()
    except ValueError:
        dt_from = today.replace(day=1)
        dt_to   = today
    bookings = Booking.objects.filter(
        date_of_event__gte=dt_from,
        date_of_event__lte=dt_to,
    ).order_by('date_of_event', 'start_time')
    return bookings, dt_from, dt_to


@login_required
def export_excel(request):
    bookings, dt_from, dt_to = _get_export_queryset(request)
    config    = SystemConfiguration.objects.first()
    brgy_name = config.barangay_name if config else "Barangay Gym Court"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"
    header_fill  = PatternFill("solid", fgColor="1A56DB")
    subhead_fill = PatternFill("solid", fgColor="DBEAFE")
    alt_fill     = PatternFill("solid", fgColor="F8FAFC")
    paid_fill    = PatternFill("solid", fgColor="DCFCE7")
    unpaid_fill  = PatternFill("solid", fgColor="FEE2E2")
    partial_fill = PatternFill("solid", fgColor="FEF9C3")
    cancel_fill  = PatternFill("solid", fgColor="F1F5F9")
    col_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    muted_font= Font(name="Calibri", size=10, color="64748B")
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    ws.merge_cells("A1:H1")
    ws["A1"] = brgy_name
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color="0F172A")
    ws["A1"].alignment = center
    ws.merge_cells("A2:H2")
    ws["A2"] = "Booking Records Export"
    ws["A2"].font      = Font(name="Calibri", bold=True, size=12, color="1A56DB")
    ws["A2"].alignment = center
    ws.merge_cells("A3:H3")
    ws["A3"] = f"Date Range: {dt_from.strftime('%B %d, %Y')} — {dt_to.strftime('%B %d, %Y')}"
    ws["A3"].font      = Font(name="Calibri", size=10, color="475569")
    ws["A3"].alignment = center
    ws.merge_cells("A4:H4")
    ws["A4"] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}   |   Total Records: {bookings.count()}"
    ws["A4"].font      = Font(name="Calibri", size=9, color="94A3B8")
    ws["A4"].alignment = center
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16
    headers    = ["Receipt #", "Event Name", "Customer Name", "Contact", "Date of Event", "Time", "Booking Status", "Payment"]
    col_widths = [16, 24, 22, 16, 18, 16, 16, 12]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font      = col_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[5].height = 20
    for row_idx, booking in enumerate(bookings, start=6):
        is_alt       = (row_idx % 2 == 0)
        is_cancelled = booking.status == 'CANCELLED'
        row_data = [
            booking.receipt_number,
            booking.event_name,
            booking.customer_name,
            booking.contact_number,
            booking.date_of_event.strftime('%b %d, %Y'),
            f"{booking.start_time.strftime('%I:%M %p')} ({booking.duration_hours}h)",
            booking.get_status_display(),
            booking.get_payment_status_display(),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = muted_font if is_cancelled else data_font
            cell.alignment = center if col_idx in [1, 5, 6, 7, 8] else left
            cell.border    = border
            if is_cancelled:
                cell.fill = cancel_fill
            elif is_alt:
                cell.fill = alt_fill
            if col_idx == 8:
                if booking.payment_status == 'PAID':
                    cell.fill = paid_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="15803D")
                elif booking.payment_status == 'UNPAID':
                    cell.fill = unpaid_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
                elif booking.payment_status == 'PARTIAL':
                    cell.fill = partial_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
        ws.row_dimensions[row_idx].height = 18
    summary_row = bookings.count() + 7
    ws.merge_cells(f"A{summary_row}:H{summary_row}")
    ws[f"A{summary_row}"] = (
        f"Total: {bookings.count()}   |   "
        f"Active: {bookings.filter(status='ACTIVE').count()}   |   "
        f"Cancelled: {bookings.filter(status='CANCELLED').count()}   |   "
        f"Paid: {bookings.filter(payment_status='PAID').count()}"
    )
    ws[f"A{summary_row}"].font      = Font(name="Calibri", bold=True, size=10, color="1E3A5F")
    ws[f"A{summary_row}"].fill      = subhead_fill
    ws[f"A{summary_row}"].alignment = left
    ws.freeze_panes = "A6"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"bookings_{dt_from.strftime('%Y%m%d')}_to_{dt_to.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_pdf(request):
    bookings, dt_from, dt_to = _get_export_queryset(request)
    config    = SystemConfiguration.objects.first()
    brgy_name = config.barangay_name if config else "Barangay Gym Court"
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm,
    )
    styles   = getSampleStyleSheet()
    elements = []
    title_style = ParagraphStyle('Title', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#0F172A'),
        spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
        fontSize=11, textColor=colors.HexColor('#1A56DB'),
        spaceAfter=2, alignment=TA_CENTER)
    info_style = ParagraphStyle('Info', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748B'),
        spaceAfter=12, alignment=TA_CENTER)
    elements.append(Paragraph(brgy_name, title_style))
    elements.append(Paragraph("Booking Records Export", sub_style))
    elements.append(Paragraph(
        f"Date Range: {dt_from.strftime('%B %d, %Y')} — {dt_to.strftime('%B %d, %Y')}  "
        f"|  Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}  "
        f"|  Total: {bookings.count()} records", info_style))
    col_headers = ["Receipt #", "Event Name", "Customer", "Contact", "Date", "Time", "Status", "Payment"]
    table_data  = [col_headers]
    for booking in bookings:
        table_data.append([
            booking.receipt_number,
            booking.event_name,
            booking.customer_name,
            booking.contact_number,
            booking.date_of_event.strftime('%b %d, %Y'),
            f"{booking.start_time.strftime('%I:%M %p')} ({booking.duration_hours}h)",
            booking.get_status_display(),
            booking.get_payment_status_display(),
        ])
    col_widths_pdf = [3*cm, 5*cm, 4.5*cm, 3.5*cm, 3*cm, 3.5*cm, 2.5*cm, 2.5*cm]
    table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#1A56DB')),
        ('TEXTCOLOR',    (0,0), (-1,0),  colors.white),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0),  9),
        ('ALIGN',        (0,0), (-1,0),  'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,0),  6),
        ('BOTTOMPADDING',(0,0), (-1,0),  6),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('ALIGN',        (0,1), (-1,-1), 'LEFT'),
        ('TOPPADDING',   (0,1), (-1,-1), 5),
        ('BOTTOMPADDING',(0,1), (-1,-1), 5),
        ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('LINEBELOW',    (0,0), (-1,0),  1,   colors.HexColor('#1A56DB')),
    ])
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0,i), (-1,i), colors.white)
        else:
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#F8FAFC'))
        booking = bookings[i-1]
        if booking.payment_status == 'PAID':
            style.add('TEXTCOLOR', (7,i), (7,i), colors.HexColor('#15803D'))
            style.add('FONTNAME',  (7,i), (7,i), 'Helvetica-Bold')
        elif booking.payment_status == 'UNPAID':
            style.add('TEXTCOLOR', (7,i), (7,i), colors.HexColor('#B91C1C'))
            style.add('FONTNAME',  (7,i), (7,i), 'Helvetica-Bold')
        elif booking.payment_status == 'PARTIAL':
            style.add('TEXTCOLOR', (7,i), (7,i), colors.HexColor('#92400E'))
            style.add('FONTNAME',  (7,i), (7,i), 'Helvetica-Bold')
        if booking.status == 'CANCELLED':
            style.add('TEXTCOLOR', (0,i), (-1,i), colors.HexColor('#94A3B8'))
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1, 0.4*cm))
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#1E3A5F'),
        backColor=colors.HexColor('#DBEAFE'),
        borderPadding=6, alignment=TA_LEFT)
    elements.append(Paragraph(
        f"<b>Summary:</b>  Total: {bookings.count()}  |  "
        f"Active: {bookings.filter(status='ACTIVE').count()}  |  "
        f"Cancelled: {bookings.filter(status='CANCELLED').count()}  |  "
        f"Paid: {bookings.filter(payment_status='PAID').count()}  |  "
        f"Unpaid: {bookings.filter(payment_status='UNPAID').count()}",
        summary_style))
    doc.build(elements)
    buffer.seek(0)
    filename = f"bookings_{dt_from.strftime('%Y%m%d')}_to_{dt_to.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def profile_settings(request):
    """
    Admin profile settings page for changing password.
    Only accessible to authenticated users.
    """
    if request.method == 'POST':
        form = AdminPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            # Update the password
            new_password = form.cleaned_data['new_password']
            request.user.set_password(new_password)
            request.user.save()
            
            # Display success message
            messages.success(request, 'Your password has been changed successfully! Please log in again.')
            
            # Redirect to login page so user logs in with new password
            return redirect('login')
    else:
        form = AdminPasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'user_info': {
            'username': request.user.username,
            'email': request.user.email,
            'full_name': request.user.get_full_name() or request.user.username,
        }
    }
    return render(request, 'core/admin_profile_settings.html', context)